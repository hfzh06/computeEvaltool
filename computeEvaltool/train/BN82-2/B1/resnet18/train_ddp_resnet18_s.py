#!/usr/bin/env python3
# train_ddp_resnet18.py

import os
import argparse
import time
import csv

import torch
import torch.nn as nn
import torch.optim as optim
import torch.distributed as dist
from torch.nn.parallel import DistributedDataParallel as DDP
from torch.utils.data import DataLoader, DistributedSampler
import torchvision.transforms as transforms
import torchvision.datasets as datasets
import torchvision.models as models

from openpyxl import Workbook, load_workbook


# ======================================================
# Args
# ======================================================
def parse_args():
    parser = argparse.ArgumentParser()
    parser.add_argument('--epochs', type=int, default=90)
    parser.add_argument('--batch-size', type=int, default=128)
    parser.add_argument('--lr', type=float, default=0.05)
    parser.add_argument('--momentum', type=float, default=0.9)
    parser.add_argument('--weight-decay', type=float, default=5e-4)
    parser.add_argument('--workers', type=int, default=4)
    parser.add_argument('--save-dir', type=str, default='/root/computeEvaltool/train/B1/resnet18/checkpoints')
    parser.add_argument('--data-root', type=str, default='/data/cifar10')
    parser.add_argument('--seed', type=int, default=42)
    parser.add_argument('--warmup-epochs', type=int, default=5)
    parser.add_argument('--clip-grad', type=float, default=5.0)
    return parser.parse_args()


# ======================================================
# Distributed Init
# ======================================================
def init_distributed_mode():
    local_rank = int(os.environ.get('LOCAL_RANK', '0'))
    world_size = int(os.environ.get('WORLD_SIZE', '1'))
    rank = int(os.environ.get('RANK', '0'))

    dist.init_process_group(backend='nccl', init_method='env://')
    torch.cuda.set_device(local_rank)

    return local_rank, rank, world_size


# ======================================================
# Data
# ======================================================
def get_dataloaders(data_root, batch_size, workers):
    normalize = transforms.Normalize((0.4914, 0.4822, 0.4465),
                                     (0.2023, 0.1994, 0.2010))

    train_transform = transforms.Compose([
        transforms.RandomCrop(32, padding=4),
        transforms.RandomHorizontalFlip(),
        transforms.ToTensor(),
        normalize,
    ])

    test_transform = transforms.Compose([
        transforms.ToTensor(),
        normalize,
    ])

    train_dataset = datasets.CIFAR10(root=data_root, train=True, download=False, transform=train_transform)
    val_dataset = datasets.CIFAR10(root=data_root, train=False, download=False, transform=test_transform)

    train_sampler = DistributedSampler(train_dataset)
    val_sampler = DistributedSampler(val_dataset, shuffle=False)

    train_loader = DataLoader(train_dataset, batch_size=batch_size, sampler=train_sampler,
                              num_workers=workers, pin_memory=True)
    val_loader = DataLoader(val_dataset, batch_size=256, sampler=val_sampler,
                            num_workers=workers, pin_memory=True)

    return train_loader, val_loader, train_sampler, val_sampler


# ======================================================
# Save LAST 3 epochs, but APPEND instead of overwrite
# ======================================================
def save_last3_append(records, results_dir):
    os.makedirs(results_dir, exist_ok=True)

    xlsx_path = os.path.join(results_dir, "training_resnet18_2.xlsx")
    csv_path = os.path.join(results_dir, "training_resnet18_2.csv")

    header = ["BatchSize", "Epoch", "Time(s)", "Throughput(samples/s)", "Loss", "Acc(%)"]

    last3 = records[-3:]  # 最后 3 行

    # ---------------- Excel (Append) ----------------
    if os.path.exists(xlsx_path):
        wb = load_workbook(xlsx_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(header)

    for row in last3:
        ws.append(row)

    wb.save(xlsx_path)

    # ---------------- CSV (Append) ----------------
    write_header = not os.path.exists(csv_path)
    with open(csv_path, "a", newline="") as f:
        writer = csv.writer(f)
        if write_header:
            writer.writerow(header)
        writer.writerows(last3)

    print(f"[INFO] Appended last 3 epochs → {xlsx_path}")


# ======================================================
# Main
# ======================================================
def main():
    args = parse_args()
    local_rank, rank, world_size = init_distributed_mode()

    torch.manual_seed(args.seed + rank)
    torch.backends.cudnn.benchmark = True

    # Model
    model = models.resnet18(num_classes=10).cuda()
    model = DDP(model, device_ids=[local_rank], output_device=local_rank)

    train_loader, val_loader, train_sampler, val_sampler = get_dataloaders(
        args.data_root, args.batch_size, args.workers
    )

    optimizer = optim.SGD(model.parameters(), lr=args.lr,
                          momentum=args.momentum, weight_decay=args.weight_decay)
    criterion = nn.CrossEntropyLoss().cuda()
    scaler = torch.cuda.amp.GradScaler(enabled=True)

    effective_max_lr = args.lr * world_size
    warmup_epochs = args.warmup_epochs
    total_epochs = args.epochs

    cosine_scheduler = torch.optim.lr_scheduler.CosineAnnealingLR(
        optimizer, T_max=max(1, total_epochs - warmup_epochs), eta_min=1e-6)

    if rank == 0:
        os.makedirs(args.save_dir, exist_ok=True)

    records = []

    # ======================================================
    # Training Loop
    # ======================================================
    for epoch in range(total_epochs):
        train_sampler.set_epoch(epoch)

        # ---- LR schedule ----
        if epoch < warmup_epochs:
            lr_this_epoch = effective_max_lr * float(epoch + 1) / warmup_epochs
            for pg in optimizer.param_groups:
                pg['lr'] = lr_this_epoch
        else:
            if epoch == warmup_epochs:
                for pg in optimizer.param_groups:
                    pg['lr'] = effective_max_lr
            cosine_scheduler.step(epoch - warmup_epochs)

        # ---- Training ----
        epoch_start = time.time()
        total_train_samples = 0

        model.train()
        for images, targets in train_loader:
            images = images.cuda()
            targets = targets.cuda()

            total_train_samples += images.size(0)

            with torch.cuda.amp.autocast():
                outputs = model(images)
                loss = criterion(outputs, targets)

            optimizer.zero_grad()
            scaler.scale(loss).backward()
            scaler.unscale_(optimizer)
            torch.nn.utils.clip_grad_norm_(model.parameters(), max_norm=args.clip_grad)
            scaler.step(optimizer)
            scaler.update()

        # ---- Validation ----
        model.eval()
        val_loss_sum = 0
        val_correct = 0
        val_total = 0

        with torch.no_grad():
            for images, targets in val_loader:
                images = images.cuda()
                targets = targets.cuda()

                with torch.cuda.amp.autocast():
                    outputs = model(images)
                    loss = criterion(outputs, targets)

                val_loss_sum += loss.item() * images.size(0)
                val_correct += (outputs.argmax(1) == targets).sum().item()
                val_total += images.size(0)

        # reduce
        t_loss = torch.tensor(val_loss_sum, device='cuda')
        t_corr = torch.tensor(val_correct, device='cuda')
        t_total = torch.tensor(val_total, device='cuda')
        t_samples = torch.tensor(total_train_samples, device='cuda')

        dist.all_reduce(t_loss)
        dist.all_reduce(t_corr)
        dist.all_reduce(t_total)
        dist.all_reduce(t_samples)

        val_loss = t_loss.item() / t_total.item()
        val_acc = t_corr.item() / t_total.item() * 100
        epoch_time = time.time() - epoch_start
        throughput = t_samples.item() / epoch_time

        if rank == 0:
            print(f"Epoch {epoch} | BS={args.batch_size} | "
                  f"time={epoch_time:.2f}s | thr={throughput:.2f} | "
                  f"loss={val_loss:.2f} | acc={val_acc:.2f}%")

            # ----------- 保存的格式 ----------- 
            records.append([
                args.batch_size,
                epoch,
                round(epoch_time, 2),
                round(throughput, 2),
                round(val_loss, 2),
                round(val_acc, 2)
            ])

    # ======================================================
    # Save LAST 3 EPOCHS (APPEND)
    # ======================================================
    if rank == 0:
        save_last3_append(records, "/root/computeEvaltool/train/B1/results")

    dist.destroy_process_group()


if __name__ == "__main__":
    main()
