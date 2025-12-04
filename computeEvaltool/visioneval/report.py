from pathlib import Path
from typing import Dict, List, Tuple
import logging

from rich.console import Console
from rich.panel import Panel
from rich.table import Table
from rich.text import Text
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import time
import random

from .benchmark import calculate_percentiles, MODEL_BATCH_SIZES

logger = logging.getLogger('vision_benchmark')

def _batch_factor(model_name: str) -> int:
    return MODEL_BATCH_SIZES.get(model_name, 1) or 1

def print_benchmark_summary(
    all_results: Dict[str, Dict[int, Tuple[List[float], List[float], int]]],
    timeout: int,
    output_dir: Path = None
):
    """打印并保存 benchmark 摘要报告
    
    Args:
        all_results: {model_name: {concurrency: (elapseds, responsetimes, total_requests)}}
    """
    console = Console(width=140)
    
    # 创建标题面板
    title = Text('Vision API Benchmark Summary Report', style='bold magenta')
    console.print(Panel(title, width=100))
    
    # 汇总统计
    total_tests = sum(len(conc_results) for conc_results in all_results.values())
    total_requests_all = sum(
        result[2] 
        for model_results in all_results.values() 
        for result in model_results.values()
    )
    
    # 基本信息表
    basic_info = Table(show_header=False, width=100, box=None)
    basic_info.add_column('Metric', style='cyan', width=35)
    basic_info.add_column('Value', style='green', width=65)
    
    basic_info.add_row('Models Tested', ', '.join(all_results.keys()))
    basic_info.add_row('Test Duration (per test)', f'{timeout} seconds')
    basic_info.add_row('Total Test Configurations', str(total_tests))
    basic_info.add_row('Total Requests (all tests)', f'{total_requests_all:,}')
    
    console.print('\n📊 Basic Information:')
    console.print(basic_info)
    
    # 详细性能指标表
    table = Table(
        title='📈 Detailed Performance Metrics by Model and Concurrency',
        show_header=True,
        header_style='bold cyan',
        border_style='blue',
        width=100,
    )
    
    table.add_column('Model', justify='left', style='cyan', width=8)
    table.add_column('Conc.', justify='right', width=8)
    table.add_column('Requests', justify='right', width=8)
    table.add_column('QPS', justify='right', width=8)
    table.add_column('RPS', justify='right', width=8)
    table.add_column('Avg Lat.(ms)', justify='right', width=10)
    # table.add_column('P50 (ms)/Q', justify='right', width=15)
    # table.add_column('P90 (ms)/Q', justify='right', width=15)
    # table.add_column('P99 (ms)/Q', justify='right', width=15)
    # table.add_column('Avg Resp.(s)', justify='right', width=13)
    
    for model_name, conc_results in all_results.items():
        factor = _batch_factor(model_name)
        for conc, (elapseds, responsetimes, requests) in sorted(conc_results.items()):
            qps = requests / timeout if timeout else 0
            rps = qps * factor
            avg_latency = (sum(elapseds) / len(elapseds)) if elapseds else float('nan')
            avg_latency_per = avg_latency / factor
            percentiles = calculate_percentiles(elapseds)
            p50 = percentiles['p50'] / factor if elapseds else float('nan')
            p90 = percentiles['p90'] / factor if elapseds else float('nan')
            p99 = percentiles['p99'] / factor if elapseds else float('nan')
            avg_resp = (sum(responsetimes) / len(responsetimes)) if responsetimes else float('nan')

            table.add_row(
                model_name,
                str(conc),
                f'{requests:,}',
                f'{qps:.2f}',
                f'{rps:.2f}',
                f'{avg_latency_per:.3f}',
                # f'{p50:.3f}',
                # f'{p90:.3f}',
                # f'{p99:.3f}',
                # f'{avg_resp:.3f}'
            )
    
    console.print('\n')
    console.print(table)
    
    # 按模型汇总的统计表
    summary_table = Table(
        title='📋 Summary by Model',
        show_header=True,
        header_style='bold yellow',
        border_style='green',
        width=120,
    )
    
    summary_table.add_column('Model', justify='left', style='cyan', width=20)
    summary_table.add_column('Avg QPS', justify='right', width=15)
    summary_table.add_column('Avg RPS', justify='right', width=15)
    summary_table.add_column('Avg Lat.(ms)', justify='right', width=20)
    summary_table.add_column('P99 Lat.(ms)', justify='right', width=20)
    summary_table.add_column('Total Requests', justify='right', width=20)
    
    for model_name, conc_results in sorted(all_results.items()):
        factor = _batch_factor(model_name)
        all_elapseds = [e for result in conc_results.values() for e in result[0]]
        total_requests = sum(result[2] for result in conc_results.values())
        total_duration = timeout * len(conc_results)
        
        avg_qps = total_requests / total_duration if total_duration > 0 else 0
        avg_rps = avg_qps * factor
        avg_latency = (sum(all_elapseds) / len(all_elapseds)) / factor if all_elapseds else 0
        p99 = calculate_percentiles(all_elapseds)['p99'] / factor if all_elapseds else 0
        
        summary_table.add_row(
            model_name,
            f'{avg_qps:.2f}',
            f'{avg_rps:.2f}',
            f'{avg_latency:.3f}',
            f'{p99:.3f}',
            f'{total_requests:,}'
        )
    
    console.print('\n')
    # console.print(summary_table)
    
    # 性能建议
    recommendations = []
    for model_name, conc_results in all_results.items():
        all_elapseds = [e for result in conc_results.values() for e in result[0]]
        avg_lat = sum(all_elapseds) / len(all_elapseds) if all_elapseds else 0
        p99 = calculate_percentiles(all_elapseds)['p99']
        
        if avg_lat > 500:
            recommendations.append(f'⚠️  {model_name}: Average latency is high ({avg_lat:.1f}ms)')
        if p99 > 1000:
            recommendations.append(f'⚠️  {model_name}: P99 latency exceeds 1s ({p99:.1f}ms)')
    
    if not recommendations:
        recommendations.append('✅ All models performance looks good!')
    
    recommend_text = Text('\n💡 Performance Recommendations:', style='bold cyan')
    console.print(recommend_text)
    for rec in recommendations:
        console.print(f'  {rec}', style='yellow')
    
    # 保存到 Excel
    if output_dir:
        _save_to_excel_new(all_results, timeout, output_dir)
    
    console.print('\n')


def _save_to_excel(all_results: Dict[str, Dict[int, Tuple]], timeout: int, output_dir: Path):
    """保存结果到 Excel 文件"""
    try:
        output_dir.mkdir(parents=True, exist_ok=True)
        
        # 准备数据
        summary_data = []
        for model_name, conc_results in sorted(all_results.items()):
            for conc, (elapseds, responsetimes, requests) in sorted(conc_results.items()):
                rps = requests / timeout if timeout > 0 else 0
                avg_lat = sum(elapseds) / len(elapseds) if elapseds else float('nan')
                perc = calculate_percentiles(elapseds)
                avg_resp = sum(responsetimes) / len(responsetimes) if responsetimes else float('nan')
                
                summary_data.append([
                    model_name,
                    conc,
                    requests,
                    rps,
                    avg_lat,
                    perc['p50'],
                    perc['p90'],
                    perc['p99'],
                    avg_resp
                ])
        
        columns = ['Model', 'Concurrency', 'Requests', 'RPS', 'Avg Lat.(ms)', 
                   'P50 (ms)', 'P90 (ms)', 'P99 (ms)', 'Avg Resp.(s)']
        df = pd.DataFrame(summary_data, columns=columns)
        
        # 数据类型转换
        for c in ['Concurrency', 'Requests']:
            df[c] = pd.to_numeric(df[c], errors='coerce').astype('Int64')
        for c in ['RPS', 'Avg Lat.(ms)', 'P50 (ms)', 'P90 (ms)', 'P99 (ms)', 'Avg Resp.(s)']:
            df[c] = pd.to_numeric(df[c], errors='coerce')
        
        excel_path = output_dir / f"vision_benchmark_multi_model.xlsx"
        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="benchmark")
            ws = writer.sheets["benchmark"]
            
            # 冻结首行
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            
            # 表头样式
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for col_idx, col_name in enumerate(df.columns, start=1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # 设置列宽
                col_letter = get_column_letter(col_idx)
                values = [str(col_name)] + [str(v) for v in df[col_name].tolist()]
                max_len = min(max(len(v) for v in values) + 2, 40)
                ws.column_dimensions[col_letter].width = max(12, max_len)
                
                # 数据对齐
                for r in range(2, ws.max_row + 1):
                    ws.cell(row=r, column=col_idx).alignment = Alignment(horizontal="right")
            
            # 数字格式
            for c in ['RPS', 'Avg Lat.(ms)', 'P50 (ms)', 'P90 (ms)', 'P99 (ms)', 'Avg Resp.(s)']:
                if c in df.columns:
                    col_idx = df.columns.get_loc(c) + 1
                    for r in range(2, ws.max_row + 1):
                        ws.cell(row=r, column=col_idx).number_format = '0.000'
        
        logger.info(f"✅ Benchmark results saved to {excel_path}")
    except Exception as e:
        logger.exception(f"❌ Failed to save Excel: {str(e)}")

def _save_to_excel_new(all_results: Dict[str, Dict[int, Tuple]], timeout: int, output_dir: Path):
    """保存结果到 Excel 文件"""
    try:
        output_dir.mkdir(parents=True, exist_ok=True)
        
        summary_data = []
        for model_name, conc_results in sorted(all_results.items()):
            factor = _batch_factor(model_name)
            for conc, (elapseds, responsetimes, requests) in sorted(conc_results.items()):
                qps = requests / timeout if timeout > 0 else 0
                rps = qps * factor
                avg_lat = sum(elapseds) / len(elapseds) if elapseds else float('nan')
                avg_lat_per = avg_lat / factor if factor else float('nan')
                perc = calculate_percentiles(elapseds)
                p50 = perc['p50'] / factor if elapseds else float('nan')
                p90 = perc['p90'] / factor if elapseds else float('nan')
                p99 = perc['p99'] / factor if elapseds else float('nan')
                avg_resp = sum(responsetimes) / len(responsetimes) if responsetimes else float('nan')
                
                summary_data.append([
                    model_name,
                    conc,
                    requests,
                    qps,
                    rps,
                    avg_lat_per,
                    p50,
                    p90,
                    p99,
                    avg_resp
                ])
        
        columns = [
            'Model', 'Concurrency', 'Requests',
            'QPS', 'RPS',
            'Avg Lat.(ms) per req', 'P50 (ms)', 'P90 (ms)', 'P99 (ms)',
            'Avg Resp.(s)'
        ]
        df = pd.DataFrame(summary_data, columns=columns)
        
        for c in ['Concurrency', 'Requests']:
            df[c] = pd.to_numeric(df[c], errors='coerce').astype('Int64')
        for c in ['QPS', 'RPS', 'Avg Lat.(ms) per req', 'P50 (ms)', 'P90 (ms)', 'P99 (ms)', 'Avg Resp.(s)']:
            df[c] = pd.to_numeric(df[c], errors='coerce')
        
        rand_id = int(time.time() - 1763905562) + random.randint(1, 9999)

        excel_path = output_dir / f"vision_benchmark_multi_model_{rand_id}.xlsx"
        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="benchmark")
            ws = writer.sheets["benchmark"]
            
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for col_idx, col_name in enumerate(df.columns, start=1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
                col_letter = get_column_letter(col_idx)
                values = [str(col_name)] + [str(v) for v in df[col_name].tolist()]
                max_len = min(max(len(v) for v in values) + 2, 40)
                ws.column_dimensions[col_letter].width = max(12, max_len)
                
                for r in range(2, ws.max_row + 1):
                    ws.cell(row=r, column=col_idx).alignment = Alignment(horizontal="right")
            
            for c in ['QPS', 'RPS', 'Avg Lat.(ms) per req', 'P50 (ms)', 'P90 (ms)', 'P99 (ms)', 'Avg Resp.(s)']:
                if c in df.columns:
                    col_idx = df.columns.get_loc(c) + 1
                    for r in range(2, ws.max_row + 1):
                        ws.cell(row=r, column=col_idx).number_format = '0.000'
        
        logger.info(f"✅ Benchmark results saved to {excel_path}")
    except Exception as e:
        logger.exception(f"❌ Failed to save Excel: {e}")