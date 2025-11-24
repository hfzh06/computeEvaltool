# the following code is largely adapted from https://github.com/lework/llm-benchmark

import numpy as np
from rich.console import Console
from rich.panel import Panel
from rich.style import Style
from rich.table import Table
from rich.text import Text

from computeEvaltool.utils.log import get_logger
from .benchmark_util import Metrics
from .db_util import PercentileMetrics
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from ..arguments import Arguments
import random
import time
logger = get_logger()


def analyze_results(all_results):
    """Analyze all test results and generate a summary report"""
    summary = []
    total_tokens = 0
    total_time = 0

    for result in all_results:
        total_metrics = result[0]
        percentile_metrics = result[1]
        percentiles = percentile_metrics[PercentileMetrics.PERCENTILES]
        try:
            concurrency = total_metrics.get(Metrics.NUMBER_OF_CONCURRENCY, 0)
            rps = total_metrics.get(Metrics.REQUEST_THROUGHPUT, 0)
            avg_latency = total_metrics.get(Metrics.AVERAGE_LATENCY, 0)
            p99_latency = percentile_metrics.get(PercentileMetrics.LATENCY)[percentiles.index('99%')]
            avg_tps = total_metrics.get(Metrics.OUTPUT_TOKEN_THROUGHPUT, 0)
            avg_ttft = total_metrics.get(Metrics.AVERAGE_TIME_TO_FIRST_TOKEN, 0)
            p99_ttft = percentile_metrics.get(PercentileMetrics.TTFT)[percentiles.index('99%')]
            success_rate = (total_metrics.get(Metrics.SUCCEED_REQUESTS, 0)
                            / total_metrics.get(Metrics.TOTAL_REQUESTS, 1)) * 100
            avg_tpot = total_metrics.get(Metrics.AVERAGE_TIME_PER_OUTPUT_TOKEN, 0)
            p99_tpot = percentile_metrics.get(PercentileMetrics.TPOT)[percentiles.index('99%')]

            # Ensure all values are valid numbers
            if any(x is None for x in [concurrency, rps, avg_latency, p99_latency, avg_tps, avg_ttft]):
                logger.warning(f'Warning: Test results for concurrency {concurrency} contain invalid data, skipped')
                continue

            # summary.append([
            #     concurrency,
            #     f'{rps:.2f}' if rps is not None else 'N/A',
            #     f'{avg_latency:.3f}' if avg_latency is not None else 'N/A',
            #     f'{p99_latency:.3f}' if p99_latency is not None else 'N/A',
            #     f'{avg_tps:.2f}' if avg_tps is not None else 'N/A',
            #     f'{avg_ttft:.3f}' if avg_ttft is not None else 'N/A',
            #     f'{success_rate:.1f}%' if success_rate is not None else 'N/A',
            #     f'{p99_ttft:.3f}' if p99_ttft is not None else 'N/A',
            #     f'{avg_tpot:.3f}' if avg_tpot is not None else 'N/A',
            #     f'{p99_tpot:.3f}' if p99_tpot is not None else 'N/A',
            # ])
            summary.append([
                int(concurrency) if concurrency is not None else 0,
                float(rps) if rps is not None else float('nan'),
                float(avg_latency) if avg_latency is not None else float('nan'),
                float(p99_latency) if p99_latency is not None else float('nan'),
                float(avg_tps) if avg_tps is not None else float('nan'),
                float(avg_ttft) if avg_ttft is not None else float('nan'),
                f'{success_rate:.1f}%' if success_rate is not None else 'N/A',  # 仅供控制台显示
                float(p99_ttft) if p99_ttft is not None else float('nan'),
                float(avg_tpot) if avg_tpot is not None else float('nan'),
                float(p99_tpot) if p99_tpot is not None else float('nan'),
            ])

            total_tokens += total_metrics.get(Metrics.AVERAGE_OUTPUT_TOKENS_PER_REQUEST, 0) * total_metrics.get(
                Metrics.SUCCEED_REQUESTS, 0)
            total_time += total_metrics.get(Metrics.TIME_TAKEN_FOR_TESTS, 0)
        except Exception as e:
            logger.warning(
                f"Warning: Error processing results for concurrency {result.get('concurrency', 'unknown')}: {str(e)}")
            continue

    if not summary:
        logger.warning('Error: No valid test result data')
        return [], 0, 0

    return summary, total_tokens, total_time


def print_summary(all_results, model_name, args: Arguments):
    """Print test results summary"""
    summary, total_tokens, total_time = analyze_results(all_results)

    if not summary:
        logger.warning('No available test result data to display')
        return

    console = Console(width=100)  # Set fixed width

    # Create title panel
    title = Text('Performance Test Summary Report', style='bold')
    console.print(Panel(title, width=60))

    # Print basic information
    basic_info = Table(show_header=False, width=60)
    basic_info.add_column('Name', style='cyan', width=25)
    basic_info.add_column('Value', style='green', width=35)

    basic_info.add_row('Model', model_name)
    basic_info.add_row('Total Generated', f'{total_tokens:,} tokens')
    basic_info.add_row('Total Test Time', f'{total_time:.2f} seconds')
    basic_info.add_row('Avg Output Rate', f'{total_tokens / total_time:.2f} tokens/sec')

    console.print('\nBasic Information:')
    console.print(basic_info)

    # Create detailed performance metrics table
    table = Table(
        title='Detailed Performance Metrics',
        show_header=True,
        header_style='bold cyan',
        border_style='blue',
        width=100,  # Set total table width
        pad_edge=False,  # Reduce edge padding
        min_width=60,  # Minimum width
    )

    # Add columns (set fixed column widths)
    table.add_column('Conc.', justify='right', style='cyan')
    table.add_column('RPS', justify='right')
    table.add_column('Avg Lat.(s)', justify='right')
    table.add_column('P99 Lat.(s)', justify='right')
    table.add_column('Gen. toks/s', justify='right')
    table.add_column('Avg TTFT(s)', justify='right')
    table.add_column('P99 TTFT(s)', justify='right')
    table.add_column('Avg TPOT(s)', justify='right')
    table.add_column('P99 TPOT(s)', justify='right')
    table.add_column('Success Rate', justify='right', style='green')

    # Add data rows
    for row in summary:
        try:
            # Set row style based on success rate
            success_rate = float(row[6].rstrip('%'))
            row_style = 'green' if success_rate >= 95 else 'yellow' if success_rate >= 80 else 'red'

            table.add_row(
                str(row[0]),  # Concurrency
                f'{float(row[1]):.2f}',  # RPS
                f'{float(row[2]):.3f}',  # Average Latency
                f'{float(row[3]):.3f}',  # P99 Latency
                f'{float(row[4]):.2f}',  # Average TPS
                f'{float(row[5]):.3f}',  # First Token Latency
                f'{float(row[7]):.3f}',  # P99 TTFT
                f'{float(row[8]):.3f}',  # Average TPOT
                f'{float(row[9]):.3f}',  # P99 TPOT
                row[6],  # Success Rate
                style=row_style)
        except ValueError as e:
            console.print(f'Warning: Error processing row data: {str(e)}', style='bold red')
            continue

    console.print('\n')
    console.print(table)
    
    # sava the summary table to an Excel file
    # try:
    #     columns = [
    #         'Conc.', 'RPS', 'Avg Lat.(s)', 'P99 Lat.(s)', 'Gen. toks/s',
    #         'Avg TTFT(s)', 'P99 TTFT(s)', 'Avg TPOT(s)', 'P99 TPOT(s)'
    #     ]
    #     summary_for_excel = [row[:6] + row[7:] for row in summary]  # Exclude Success Rate for Excel
    #     df = pd.DataFrame(summary_for_excel, columns=columns)
    #     excel_path = f"{model_name}_benchmark_summary.xlsx"
    #     df.to_excel(excel_path, index=False)
    #     logger.info(f"Summary table saved to {excel_path}")
    # except Exception as e:
    #     logger.warning(f"Failed to save summary to Excel: {str(e)}")
    
    try:
        import pandas as pd
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Alignment

        columns = [
            'Conc.', 'RPS', 'Avg Lat.(s)', 'P99 Lat.(s)', 'Gen. toks/s',
            'Avg TTFT(s)', 'P99 TTFT(s)', 'Avg TPOT(s)', 'P99 TPOT(s)'
        ]
        # 去掉第7列（Success Rate），只保留前6列和后3列
        summary_for_excel = [row[:6] + row[7:] for row in summary]
        df = pd.DataFrame(summary_for_excel, columns=columns)

        # 数值化（Conc. 整数，其他转 float）
        df['Conc.'] = pd.to_numeric(df['Conc.'], errors='coerce').astype('Int64')
        for c in columns[1:]:
            df[c] = pd.to_numeric(df[c], errors='coerce')

        rand_id = int(time.time() - 1763905562) + random.randint(1, 9999)
        excel_path = f"{model_name}_{args.inference_engine}_node{args.node_num}_gpu_{args.gpu_num}_tp{args.tp_size}_dp{args.dp_size}_{rand_id}_benchmark_summary.xlsx"
        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="summary")
            ws = writer.sheets["summary"]

            # 冻结首行 + 自动筛选
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

            # 自动列宽 + 右对齐
            for col_idx, col_name in enumerate(df.columns, start=1):
                col_letter = get_column_letter(col_idx)
                values = [str(col_name)] + [str(v) for v in df[col_name].tolist()]
                max_len = min(max(len(v) for v in values) + 2, 36)
                ws.column_dimensions[col_letter].width = max(8, max_len)
                for r in range(2, ws.max_row + 1):
                    ws.cell(row=r, column=col_idx).alignment = Alignment(horizontal="right")

            # 数字显示格式：Conc. 无小数，其余四位小数
            def fmt_col(name, numfmt):
                if name not in df.columns:
                    return
                c = df.columns.get_loc(name) + 1
                for r in range(2, ws.max_row + 1):
                    ws.cell(row=r, column=c).number_format = numfmt

            fmt_col('Conc.', '0')
            for c in columns[1:]:
                fmt_col(c, '0.0000')
                
            ws['J1'] = 'Total Time'
            ws['J2'] = total_time

        logger.info(f"Summary table saved to {excel_path}")
    except Exception as e:
        logger.warning(f"Failed to save summary to Excel: {str(e)}")


    # Calculate and display best performance configuration
    try:
        best_rps_idx = np.argmax([float(row[1]) if row[1] != 'N/A' else -1 for row in summary])
        best_latency_idx = np.argmin([float(row[2]) if row[2] != 'N/A' else float('inf') for row in summary])

        perf_info = Table(title='Best Performance Configuration', show_header=False, box=None, width=60)
        perf_info.add_column('Metric', style='cyan', width=20)
        perf_info.add_column('Value', style='green', width=40)

        perf_info.add_row('Highest RPS', f'Concurrency {summary[best_rps_idx][0]} ({summary[best_rps_idx][1]} req/sec)')
        perf_info.add_row('Lowest Latency',
                          f'Concurrency {summary[best_latency_idx][0]} ({summary[best_latency_idx][2]} seconds)')

        console.print('\n')
        console.print(perf_info)

        # Performance recommendations
        recommendations = []
        if best_rps_idx == len(summary) - 1:
            recommendations.append(
                'The system seems not to have reached its performance bottleneck, try higher concurrency')
        elif best_rps_idx == 0:
            recommendations.append('Consider lowering concurrency, current load may be too high')
        else:
            recommendations.append(f'Optimal concurrency range is around {summary[best_rps_idx][0]}')

        success_rate = float(summary[-1][6][:-1])
        if success_rate < 95:
            recommendations.append(
                'Success rate is low at high concurrency, check system resources or reduce concurrency')

        recommend_text = Text('\nPerformance Recommendations:', style='bold cyan')
        console.print(recommend_text)
        for rec in recommendations:
            console.print(f'• {rec}', style='yellow')

    except Exception as e:
        console.print(f'Warning: Error generating performance analysis: {str(e)}', style='bold red')
